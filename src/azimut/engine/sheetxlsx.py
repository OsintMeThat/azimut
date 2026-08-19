"""Read a workbook into one table per tab, so a binder arrives whole.

The three binders this tool was written against have five, six and one tab. Two of the
six hold nothing but pasted screenshots. Asking the analyst to open each one and export
it to CSV by hand is asking them not to bother, and the tab that gets skipped is always
the "How to use" one that says what the others mean.

**Reading only, and into the same tables every other import lands in.** The case's
artifact is a CSV — that is the whole design (`engine/sheets.py`) — so nothing here
writes a workbook, and every tab goes through `sheets.normalize` like a dropped file.
What comes out is a list of tables, and it is the caller that decides which of them to
file: a tab of screenshots reads as empty, and filing an empty sheet nobody asked for is
worse than saying it was empty.

**A cell becomes the words a person would read.** A workbook has no words — it has a
serial number and a display format — so this is the one importer that has to choose a
spelling. It chooses **ISO 8601**, because that is the one spelling that means the same
thing in every locale, `parse_when` reads it, and it round-trips through a spreadsheet
unchanged. A column of `dd/MM/yyyy` is one cleaning pass away for anyone who wants it.

**Formulas arrive as their last answer, not as their text.** That is what the analyst was
looking at, and `=VLOOKUP(...)` in a cell is a reference to a workbook the case will not
have. A formula Excel never cached reads as empty, which is honest: nothing computed it.
"""

from __future__ import annotations

import io
import zipfile
from datetime import date, datetime, time
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from .sheets import MAX_COLUMNS, MAX_ROWS, SheetError

#: How many tabs one workbook may be read into sheets. Six is the biggest binder seen;
#: past a couple of dozen the file is a database export and the grid is the wrong tool.
MAX_TABS = 24

#: How large a workbook may be before it is refused unread. An .xlsx is zipped XML, so
#: this is several times the CSV bound it expands into.
MAX_XLSX_BYTES = 32 * 1024 * 1024

#: How far a workbook may expand once unzipped. An .xlsx is compressed XML and a real one
#: expands ten or fifteen times over, so this leaves a full-size workbook room while
#: refusing the shape it is here for: thirty two megabytes of zeroes declaring gigabytes of
#: `sharedStrings.xml`, which `read_only` streams no part of. The app already bounds images
#: for exactly this reason (`MAX_IMAGE_PIXELS`), and an archive from a third party is the
#: same class of file — the binders this reader exists for arrive by email.
MAX_XLSX_UNZIPPED_BYTES = 512 * 1024 * 1024


def _clock(value: datetime | time) -> str:
    """The time of day a cell holds, to the second when it has one.

    The seconds are written because ISO 8601 carries them and `parse_when` reads them: a
    workbook's timecode column is what an offset is counted in, and cutting `00:01:23` to
    `00:01` moved a synchronised video by twenty three seconds with nothing said. They are
    left off when they are zero, so a column of days and hours keeps the spelling somebody
    typed rather than gaining `:00` on every row.

    Midnight is empty, which is how a `datetime` with no time in it reads as a date.
    """
    if not (value.hour or value.minute or value.second):
        return ""
    return value.strftime("%H:%M:%S" if value.second else "%H:%M")


def _cell(value: Any) -> str:
    """One cell as the text a reader would see.

    Booleans come back as the two words a `boolean` column is born with, so a tick column
    imported out of Excel and one typed into the grid hold the same file. A float that is
    a whole number loses its `.0`: Excel stores every number as a double, and a column of
    counts reading `12.0` is a column nobody would have typed.
    """
    if value is None:
        return ""
    if isinstance(value, bool):
        return "YES" if value else "NO"
    if isinstance(value, datetime):
        stamp = value.strftime("%Y-%m-%d")
        clock = _clock(value)
        return stamp if not clock else f"{stamp} {clock}"
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return _clock(value) or "00:00"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def _trim(rows: list[list[str]]) -> list[list[str]]:
    """Drop the trailing empty rows and columns a spreadsheet always reports.

    Excel's idea of how far a sheet extends outlives whatever was deleted from it, so a
    tab of eleven rows arrives claiming nine hundred and sixty-nine. The binders' own
    timeline tab does exactly that.
    """
    while rows and not any(cell.strip() for cell in rows[-1]):
        rows.pop()
    if not rows:
        return []
    width = max(
        (index + 1 for row in rows for index, cell in enumerate(row) if cell.strip()),
        default=0,
    )
    return [row[:width] + [""] * max(0, width - len(row)) for row in rows]


def _unzipped_size(data: bytes) -> int:
    """What the archive says it expands into, read off the directory and not by unzipping.

    `zipfile` stops each member at its declared size and checks the CRC, so the declaration
    is the bound the reader will honour: refusing on it costs no decompression at all.
    """
    try:
        with zipfile.ZipFile(io.BytesIO(data)) as archive:
            return sum(max(0, entry.file_size) for entry in archive.infolist())
    except (zipfile.BadZipFile, OSError, ValueError):
        # Not a zip, which `load_workbook` is about to say in the sentence written for it.
        return 0


def read_book(data: bytes) -> dict[str, Any]:
    """Every tab of a workbook, as a heading row and a body, plus what was left out.

    Answers all of them, empty ones included, each with its own name: which tabs are
    worth filing is the caller's question, and a reader that silently dropped the empty
    ones would hide the two tabs of pasted screenshots that a timeline binder keeps its
    proof of the hour in.

    Three ceilings hold — the tabs, the rows and the columns — and each of them is
    **answered rather than applied in silence**. A thirty thousand row export arriving as
    twenty thousand under a toast reading "5 sheets created" is a sheet that looks whole,
    and there is nothing on screen to suspect otherwise. So ``dropped`` names the tabs past
    `MAX_TABS` and each tab says whether it was cut, which is what the caller says out loud.
    """
    if len(data) > MAX_XLSX_BYTES:
        raise SheetError(f"a workbook is read up to {MAX_XLSX_BYTES // (1024 * 1024)} MB")
    if _unzipped_size(data) > MAX_XLSX_UNZIPPED_BYTES:
        raise SheetError(
            "a workbook is read up to "
            f"{MAX_XLSX_UNZIPPED_BYTES // (1024 * 1024)} MB once unzipped"
        )
    try:
        book = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except (
        # An .xlsx is a zip of XML, so a mis-picked file fails at any of three layers:
        # not a zip at all, a zip without the parts a workbook has, or XML that does not
        # parse. All three are the same answer to the analyst.
        zipfile.BadZipFile,
        InvalidFileException,
        OSError,
        ValueError,
        KeyError,
    ) as exc:
        raise SheetError(f"this file is not readable as a workbook: {exc}") from exc

    tables: list[dict[str, Any]] = []
    dropped = [str(tab.title) for tab in book.worksheets[MAX_TABS:]]
    try:
        for tab in book.worksheets[:MAX_TABS]:
            rows: list[list[str]] = []
            over_row = False
            over_column = False
            # One row and one column past each ceiling, read only to know whether the tab
            # goes further. What lies past them is never kept, so the extra pair costs a
            # cell each and buys the sentence that says the tab was cut.
            for line in tab.iter_rows(
                max_row=MAX_ROWS + 2, max_col=MAX_COLUMNS + 1, values_only=True
            ):
                cells = [_cell(value) for value in line]
                if any(cell.strip() for cell in cells[MAX_COLUMNS:]):
                    over_column = True
                if len(rows) > MAX_ROWS:
                    over_row = over_row or any(cell.strip() for cell in cells[:MAX_COLUMNS])
                    continue
                rows.append(cells[:MAX_COLUMNS])
            rows = _trim(rows)
            tables.append(
                {
                    "title": str(tab.title or "Sheet"),
                    "columns": rows[0] if rows else [],
                    "rows": rows[1:] if len(rows) > 1 else [],
                    #: How many were kept, when the tab held more. `None` is a tab that
                    #: fitted, which is every tab a binder has.
                    "cut_rows": MAX_ROWS if over_row else None,
                    "cut_columns": MAX_COLUMNS if over_column else None,
                }
            )
    finally:
        book.close()
    return {"tabs": tables, "dropped": dropped}
